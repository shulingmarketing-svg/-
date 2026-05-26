import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_HEADER  = "1B3A6B"
C_SUB     = "2E5F8A"
C_MISO    = "FFF0D6"   # 味噌 - 暖米色
C_MIRIN   = "D6EAD6"   # 味醂 - 淺綠
C_VINEGAR = "D6E8F5"   # 烏醋 - 淺藍
C_RICE_V  = "FDE8D8"   # 米醋 - 淺橘
C_CHILI   = "FDDADA"   # 辣椒醬 - 淺紅
C_RULE_Y  = "FFFACD"   # 規範 - 淺黃
C_RULE_R  = "FFE4E4"   # 禁止 - 淺紅
C_RULE_G  = "E4FFE4"   # 鼓勵 - 淺綠
C_LIGHT   = "F7F7F7"
C_WHITE   = "FFFFFF"

thin  = Side(style="thin",  color="CCCCCC")
thick = Side(style="medium", color="888888")
border     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
border_top = Border(left=thin,  right=thin,  top=thick, bottom=thin)

def hdr(ws, row, col, val, bg=C_HEADER, color="FFFFFF", size=11, bold=True, center=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, color=color, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    h = "center" if center else "left"
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=True)
    cell.border = border
    return cell

def data(ws, row, col, val, bg=C_WHITE, bold=False, center=False, size=10, h_height=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    h = "center" if center else "left"
    cell.alignment = Alignment(horizontal=h, vertical="top", wrap_text=True)
    cell.border = border
    return cell

def set_col(ws, cols_widths):
    for col, w in cols_widths.items():
        ws.column_dimensions[col].width = w

def merge_hdr(ws, r1, c1, r2, c2, val, bg=C_HEADER, color="FFFFFF", size=11):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font = Font(bold=True, color=color, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

wb = openpyxl.Workbook()

# ════════════════════════════════════════
# Sheet 1：KOL 個人資料卡（可複製填寫）
# ════════════════════════════════════════
ws1 = wb.active
ws1.title = "①KOL個人資料卡"

merge_hdr(ws1, 1,1, 1,4, "十全特好 × KOL 合作資料卡（一人一份）", C_HEADER, size=13)
ws1.row_dimensions[1].height = 30

fields = [
    ("KOL 姓名 / 帳號名稱", ""),
    ("合作平台", "□ IG   □ FB   □ YouTube   □ 其他：___"),
    ("主打商品", "□ 味噌   □ 味醂   □ 烏醋   □ 米醋   □ 辣椒醬"),
    ("專屬折扣碼", "（請填入，例：COCO10）"),
    ("專屬 UTM 連結", "（請填入完整連結）"),
    ("建議發文時間", ""),
    ("合作窗口聯絡人", ""),
    ("窗口 LINE / Email", ""),
    ("", ""),
    ("發文件數", "□ 1 篇   □ 2 篇   □ 其他：___"),
    ("發文格式", "□ Reels   □ 靜態貼文   □ 輪播   □ Stories"),
    ("截稿確認日", "（KOL 發文前需送審）"),
    ("正式發文日", ""),
    ("成效回報日", "發文後第 7 天"),
    ("", ""),
    ("備注事項", ""),
]

for i, (label, placeholder) in enumerate(fields, 2):
    bg = C_LIGHT if i % 2 == 0 else C_WHITE
    if not label:
        ws1.row_dimensions[i].height = 8
        continue
    hdr(ws1, i, 1, label, bg=C_SUB, size=10)
    cell = ws1.cell(row=i, column=2, value=placeholder)
    cell.font = Font(size=10, color="888888" if placeholder else "000000")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    ws1.row_dimensions[i].height = 22

set_col(ws1, {"A":22, "B":20, "C":20, "D":20})
ws1.freeze_panes = "A2"

# 注意事項區
r = len(fields) + 3
merge_hdr(ws1, r, 1, r, 4, "⚠️ 發文規範（請 KOL 務必閱讀）", "C62828")
ws1.row_dimensions[r].height = 22
r += 1

rules = [
    ("✅ 必須包含", "折扣碼、官網連結、@十全特好、指定 Hashtag"),
    ("✅ 必須確認", "發文前送審，通過後才正式發出"),
    ("❌ 不可使用", "競品品牌名稱、療效聲稱、「最頂級」「保證」等誇大詞"),
    ("❌ 健康聲稱", "不得宣稱「治療」「改善疾病」「提升免疫力」等未經認證說法"),
    ("📸 視覺要求", "自然光拍攝、生活感、商品自然入鏡（非強迫正面露出）"),
]
for label, content in rules:
    bg = C_RULE_G if "✅" in label else C_RULE_R if "❌" in label else C_RULE_Y
    data(ws1, r, 1, label, bg=bg, bold=True, center=True)
    cell = ws1.cell(row=r, column=2, value=content)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws1.row_dimensions[r].height = 30
    r += 1

# ════════════════════════════════════════
# Sheet 2：商品賣點卡
# ════════════════════════════════════════
ws2 = wb.create_sheet("②商品賣點卡")
merge_hdr(ws2, 1, 1, 1, 5, "商品賣點卡｜KOL 說話參考（請依主打商品取用對應欄）", C_HEADER, size=13)
ws2.row_dimensions[1].height = 30

products = [
    ("味噌", C_MISO, [
        ("成分單純", "「黃豆、米麴、鹽，就這三樣。」"),
        ("天然釀造", "「讓菌自己慢慢發酵，不催熟。」"),
        ("用途廣", "「不只湯，醃肉、醬汁、拌麵都可以。」"),
        ("差異化", "「快速製程的味噌，你的舌頭會告訴你差在哪。」"),
    ]),
    ("味醂", C_MIRIN, [
        ("不是米酒", "「去腥 + 增甜 + 讓食材有光澤，三件事一次。」"),
        ("照燒關鍵", "「那個油亮亮的感覺，是味醂，不是糖。」"),
        ("天然釀造", "「用糯米釀的，不是調出來的甜。」"),
        ("建議用量", "「替代米酒的時候，量減半，它比較甜。」"),
    ]),
    ("烏醋", C_VINEGAR, [
        ("無焦糖色素", "「顏色深是原料本身，不是加色素加出來的。」"),
        ("釀造基底", "「翻一下成分表，沒有你看不懂的字。」"),
        ("用途廣", "「炒菜、滷味、沾醬，台灣料理的底。」"),
        ("對比說法", "「不是每瓶烏醋都一樣，只是大部分人沒翻過成分表。」"),
    ]),
    ("米醋", C_RICE_V, [
        ("酸味圓潤", "「比白醋溫和，不嗆，涼拌最好用。」"),
        ("用途廣", "「涼拌、壽司飯、醃漬，夏天的必備。」"),
        ("天然釀造", "「用米釀的，酸味裡有穀物的甜。」"),
        ("季節感", "「夏天冰一碗米醋涼拌，比冷氣還涼。」"),
    ]),
    ("辣椒醬", C_CHILI, [
        ("成分乾淨", "「辣椒、蒜、鹽，沒有多的東西。」"),
        ("天然辣", "「天然辣和合成辣在嘴裡感覺不一樣，試過才知道。」"),
        ("萬用", "「沾、炒、拌、醃，一瓶用四種方式。」"),
        ("喉嚨感", "「吃完喉嚨不燒，是天然辣的特徵。」"),
    ]),
]

hdr(ws2, 2, 1, "商品", bg=C_SUB, size=10)
hdr(ws2, 2, 2, "賣點主題", bg=C_SUB, size=10)
hdr(ws2, 2, 3, "建議說話方式（可直接引用）", bg=C_SUB, size=10)
hdr(ws2, 2, 4, "搭配食譜方向", bg=C_SUB, size=10)
hdr(ws2, 2, 5, "備注 / KOL 自訂", bg=C_SUB, size=10)
ws2.row_dimensions[2].height = 22

recipe_hint = {
    "味噌": "味噌湯、味噌燒鮭魚、味噌炒高麗菜",
    "味醂": "日式照燒雞、滷五花肉、煮魚",
    "烏醋": "醋溜高麗菜、糖醋排骨、烏醋麵線",
    "米醋": "涼拌小黃瓜、壽司醋飯、快速醃漬小菜",
    "辣椒醬": "辣炒蛤蜊、辣拌麵、辣醃烤雞腿",
}

r = 3
for prod, bg, points in products:
    start_r = r
    for i, (theme, script) in enumerate(points):
        data(ws2, r, 1, prod if i == 0 else "", bg=bg, bold=(i==0), center=True)
        data(ws2, r, 2, theme, bg=bg, bold=True)
        data(ws2, r, 3, script, bg=bg)
        data(ws2, r, 4, recipe_hint[prod] if i == 0 else "", bg=bg)
        data(ws2, r, 5, "", bg=C_WHITE)
        ws2.row_dimensions[r].height = 28
        r += 1
    if len(points) > 1:
        ws2.merge_cells(start_row=start_r, start_column=1,
                        end_row=start_r+len(points)-1, end_column=1)
        ws2.merge_cells(start_row=start_r, start_column=4,
                        end_row=start_r+len(points)-1, end_column=4)

set_col(ws2, {"A":10, "B":18, "C":45, "D":28, "E":20})
ws2.freeze_panes = "A3"

# ════════════════════════════════════════
# Sheet 3：食譜建議
# ════════════════════════════════════════
ws3 = wb.create_sheet("③食譜建議")
merge_hdr(ws3, 1, 1, 1, 6, "食譜建議｜每種商品 3 道，KOL 可自由選擇 1 道拍攝", C_HEADER, size=13)
ws3.row_dimensions[1].height = 30

for ci, h in enumerate(["商品","食譜名稱","食材","重點步驟","拍攝建議","適合平台"], 1):
    hdr(ws3, 2, ci, h, bg=C_SUB, size=10)
ws3.row_dimensions[2].height = 22

recipes = [
    ("味噌", C_MISO, [
        ("味噌豬肉湯",
         "豬肉片、白蘿蔔、豆腐、蔥、十全味噌",
         "1. 蘿蔔先煮軟\n2. 加豬肉片燙熟\n3. 關小火，味噌用篩網濾入\n4. 不要再煮滾，蔥花起鍋",
         "熱湯冒煙感，秋天橘調光線，木碗或陶碗",
         "IG Reels / FB 靜態"),
        ("味噌燒鮭魚",
         "鮭魚排、十全味噌、味醂、少許糖",
         "1. 味噌 + 味醂 + 糖調成醬\n2. 鮭魚醃 30 分鐘\n3. 中小火煎，不要翻太多次\n4. 皮朝下先煎到酥",
         "煎到油亮的鮭魚特寫，日式餐具搭配",
         "IG 靜態 / 輪播"),
        ("味噌炒高麗菜",
         "高麗菜、蒜、十全味噌、少許水",
         "1. 味噌先用水化開備用\n2. 大火爆香蒜末\n3. 高麗菜下鍋大火炒\n4. 最後淋入味噌水，快炒 30 秒",
         "炒鍋翻炒動態，自然光，鍋邊有些焦感",
         "IG Reels（快剪）"),
    ]),
    ("味醂", C_MIRIN, [
        ("日式照燒雞腿",
         "雞腿排、醬油 2 匙、十全味醂 2 匙、糖 1 匙",
         "1. 醬汁調好備用\n2. 雞腿皮朝下中火煎出油\n3. 倒入醬汁小火收汁\n4. 反覆淋醬讓表面油亮",
         "油亮光澤特寫，醬汁在鍋裡起泡的瞬間",
         "IG Reels / FB"),
        ("滷五花肉",
         "五花肉、醬油、十全味醂、冰糖、蒜、八角",
         "1. 五花肉切塊煎至表面微焦\n2. 加醬油、味醂、糖、蒜、八角\n3. 加水蓋過食材，小火燉 50 分鐘\n4. 開蓋收汁至濃稠",
         "紅燒色澤，油亮表面，可搭配白飯入鏡",
         "IG 靜態 / 輪播步驟"),
        ("日式煮魚",
         "白肉魚（赤魗/鯖魚）、十全味醂、醬油、薑",
         "1. 鍋中放味醂先煮 1 分鐘讓酒精揮發\n2. 加醬油、薑片\n3. 魚放入，湯汁中火煮 8 分鐘\n4. 起鍋淋上剩餘湯汁",
         "日式陶碗盛裝，薑絲點綴，湯汁清亮感",
         "IG 靜態"),
    ]),
    ("烏醋", C_VINEGAR, [
        ("醋溜高麗菜",
         "高麗菜、蒜、辣椒、十全烏醋",
         "1. 大火爆香蒜末辣椒\n2. 高麗菜下鍋大火快炒\n3. 加少許鹽\n4. 烏醋最後下，快炒 10 秒起鍋",
         "大火炒菜動態，高麗菜帶一點焦邊",
         "IG Reels（15 秒快剪）"),
        ("糖醋排骨",
         "小排、十全烏醋、番茄醬、糖、蒜",
         "1. 小排炸至金黃\n2. 蒜末爆香，加番茄醬、糖、烏醋炒成醬\n3. 排骨回鍋裹醬\n4. 撒白芝麻",
         "紅亮的排骨特寫，裹滿醬汁感",
         "IG 靜態 / FB"),
        ("烏醋麵線",
         "麵線、大腸或豆干、十全烏醋、蒜泥、香菜",
         "1. 麵線燙熟備用\n2. 大腸（豆干）切片\n3. 加蒜泥、烏醋、少許醬油拌勻\n4. 香菜點綴",
         "台味十足，陶碗盛裝，烏醋色澤深邃",
         "IG 靜態 / FB"),
    ]),
    ("米醋", C_RICE_V, [
        ("涼拌小黃瓜",
         "小黃瓜、蒜末、辣椒、十全米醋、鹽、少許糖",
         "1. 小黃瓜拍碎，不要切\n2. 加蒜末、辣椒、米醋、鹽、糖\n3. 拌勻後冰 20 分鐘再吃\n4. 吃前再拌一次",
         "清爽感，白色盤子，自然光俯拍",
         "IG 靜態（夏天應景）"),
        ("壽司醋飯",
         "白飯 2 碗、十全米醋 3 匙、糖 2 匙、鹽 1 匙",
         "1. 糖鹽先溶於米醋（微溫加速）\n2. 白飯趁熱倒入壽司桶或大碗\n3. 淋入醋液，用飯匙切拌（不要壓）\n4. 邊拌邊用扇子搖涼",
         "壽司飯蓬鬆感，木飯桶或大碗，俯拍",
         "IG 輪播（步驟圖）"),
        ("快速醃漬小菜",
         "紅蘿蔔、小黃瓜、十全米醋、糖、鹽",
         "1. 糖鹽先溶解於米醋（冷的）\n2. 蔬菜切薄片\n3. 醋液倒入蔬菜\n4. 冰一晚最入味",
         "玻璃罐裝，色彩繽紛，質感好拍",
         "IG 靜態 / 輪播"),
    ]),
    ("辣椒醬", C_CHILI, [
        ("辣炒蛤蜊",
         "蛤蜊、十全辣椒醬、薑、蒜、九層塔",
         "1. 大火爆香薑蒜\n2. 辣椒醬下鍋炒香\n3. 蛤蜊下鍋，蓋鍋蓋\n4. 開口馬上起鍋，九層塔拌入",
         "蛤蜊開口瞬間，九層塔點綴，蒸氣感",
         "IG Reels（15 秒，開口瞬間是亮點）"),
        ("辣拌麵",
         "麵條、十全辣椒醬、麻油、蔥花、醬油少許",
         "1. 辣椒醬 + 麻油 + 醬油先拌勻在碗底\n2. 熱麵直接倒入碗中\n3. 拌勻讓醬汁裹上麵條\n4. 蔥花、白芝麻點綴",
         "深夜食堂感，深色碗，油光感，俯拍",
         "IG Reels / FB"),
        ("辣醃烤雞腿",
         "雞腿排、十全辣椒醬、蒜末、醬油、少許糖",
         "1. 雞腿劃刀讓醃料入味\n2. 辣椒醬 + 蒜末 + 醬油 + 糖抹勻\n3. 冰一晚\n4. 烤箱 200 度 25 分鐘，中途翻面",
         "烤到微焦的雞腿皮，烤盤自然感",
         "IG 靜態 / 輪播（醃→烤→成品）"),
    ]),
]

r = 3
for prod, bg, rlist in recipes:
    for i, (name, ingredients, steps, photo, platform) in enumerate(rlist):
        data(ws3, r, 1, prod if i==0 else "", bg=bg, bold=(i==0), center=True)
        data(ws3, r, 2, name, bg=bg, bold=True)
        data(ws3, r, 3, ingredients, bg=bg)
        data(ws3, r, 4, steps, bg=bg)
        data(ws3, r, 5, photo, bg=bg)
        data(ws3, r, 6, platform, bg=bg, center=True)
        ws3.row_dimensions[r].height = 90
        r += 1
    if len(rlist) > 1:
        ws3.merge_cells(start_row=r-len(rlist), start_column=1,
                        end_row=r-1, end_column=1)

set_col(ws3, {"A":10, "B":18, "C":28, "D":40, "E":30, "F":18})
ws3.freeze_panes = "A3"

# ════════════════════════════════════════
# Sheet 4：CTA 話術 + 禁用詞
# ════════════════════════════════════════
ws4 = wb.create_sheet("④CTA話術+語調規範")
merge_hdr(ws4, 1, 1, 1, 3, "CTA 話術模板 + 品牌語調規範", C_HEADER, size=13)
ws4.row_dimensions[1].height = 30

# CTA 話術
merge_hdr(ws4, 2, 1, 2, 3, "▍ CTA 話術模板（依主打商品選用，填入折扣碼後直接使用）", C_SUB, size=10)
ws4.row_dimensions[2].height = 22

cta_headers = ["主打商品", "建議 CTA 話術", "備注"]
for ci, h in enumerate(cta_headers, 1):
    hdr(ws4, 3, ci, h, bg=C_SUB, size=10)
ws4.row_dimensions[3].height = 22

ctas = [
    ("味噌",
     "最近煮日式料理都在用十全的味噌，成分只有黃豆、米麴、鹽，跟外面很多加一堆東西的不一樣。他們現在有在做抽獎活動，去官網加入會員就有資格，我有專屬代碼【折扣碼】可以拿 88 元折扣券，連結放在留言區／限動。",
     "「連結在哪」依平台調整（IG 放 bio 或限動、FB 放留言）"),
    ("味醂",
     "一直有人問我做照燒那個亮亮的感覺怎麼來的，其實就是味醂，我用十全的。天然釀造，不是調出來的甜。現在去他們官網加入會員可以抽獎，用我的代碼【折扣碼】首購還有 88 元折扣，連結放在下面。",
     "可搭配照燒料理實作影片"),
    ("烏醋",
     "我翻過很多品牌的烏醋成分表，十全是少數沒有焦糖色素的。顏色深是原料本身，不是加出來的。他們官網現在有抽獎活動，加入會員就有資格，加入還送 88 元折扣券，代碼是【折扣碼】，連結在留言。",
     "可搭配成分表翻書畫面"),
    ("米醋",
     "做涼拌、醃漬我現在固定用十全米醋，酸味比較圓，不會太嗆。他們官網會員可以參加抽獎活動，用我的代碼【折扣碼】加入還有 88 元折扣，給你們試試看，連結在下方。",
     "夏天發文效果最好"),
    ("辣椒醬",
     "十全辣椒醬我最近拿來炒蛤蜊、拌麵都很好用，成分是辣椒、蒜、鹽，沒有多餘的東西。官網現在可以加入會員參加抽獎，三個月共抽 900 名，我的代碼【折扣碼】加入有 88 元折扣，連結放留言區。",
     "強調成分時可搭配成分表畫面"),
]

for i, (prod, script, note) in enumerate(ctas, 4):
    bg = C_LIGHT if i%2==0 else C_WHITE
    data(ws4, i, 1, prod, bg=bg, bold=True, center=True)
    data(ws4, i, 2, script, bg=bg)
    data(ws4, i, 3, note, bg=bg)
    ws4.row_dimensions[i].height = 80

r = 4 + len(ctas) + 1

# 語調規範
merge_hdr(ws4, r, 1, r, 3, "▍ 品牌語調規範", C_SUB, size=10)
ws4.row_dimensions[r].height = 22
r += 1

hdr(ws4, r, 1, "類別", bg=C_SUB, size=10)
hdr(ws4, r, 2, "內容", bg=C_SUB, size=10)
hdr(ws4, r, 3, "範例", bg=C_SUB, size=10)
ws4.row_dimensions[r].height = 22
r += 1

rules = [
    ("✅ 鼓勵說法", C_RULE_G,
     "「成分看得懂」「釀造」「原料單純」「試過才知道」「比你想的好用」「拿起來聞」",
     "「成分只有三樣，你都看得懂。」"),
    ("⚠️ 小心使用", C_RULE_Y,
     "「天然」要有成分佐證；「無添加」需說明是哪類；「純釀造」需符合十全品項說明",
     "「天然釀造的味噌（黃豆、米麴、鹽）」✅"),
    ("❌ 絕對禁止", C_RULE_R,
     "「最頂級」「市場唯一」「保證效果」「治療」「改善疾病」「提升免疫力」「第一名」競品品牌名稱",
     "「吃了可以改善腸道健康」❌"),
    ("📝 整體語氣", C_LIGHT,
     "像朋友分享，不像業配。說「我用過」「我試過」，有個人觀點，不要全程誇讚",
     "「老實說一開始不確定，試了兩次才愛上。」"),
]

for cat, bg, content, ex in rules:
    data(ws4, r, 1, cat, bg=bg, bold=True, center=True)
    data(ws4, r, 2, content, bg=bg)
    data(ws4, r, 3, ex, bg=bg)
    ws4.row_dimensions[r].height = 50
    r += 1

set_col(ws4, {"A":14, "B":55, "C":30})
ws4.freeze_panes = "A4"

# ════════════════════════════════════════
# Sheet 5：拍攝視覺指引
# ════════════════════════════════════════
ws5 = wb.create_sheet("⑤拍攝視覺指引")
merge_hdr(ws5, 1, 1, 1, 4, "KOL 拍攝視覺指引（Olivia 出品）", C_HEADER, size=13)
ws5.row_dimensions[1].height = 30

sections = [
    ("整體風格", C_MISO, [
        ("✅ 要", "自然光（窗邊）、木桌/磁磚/麻布背景、食材有點散亂有煮過的感覺、有蒸氣有油光、商品放畫面角落或中景"),
        ("❌ 不要", "閃光燈/棚燈（假感覺）、純白無影背景（太像型錄）、擺盤太精緻像米其林、商品強迫正面露出"),
    ]),
    ("商品入鏡方式", C_MIRIN, [
        ("✅ 建議做法", "側放在料理旁邊自然露出\n烹飪時「正在倒入鍋裡」的動態\n成品後方略模糊的後景"),
        ("❌ 避免", "瓶身正對鏡頭像在拍商品照\n貼著 Logo 面向鏡頭強迫露出\n手持瓶身舉在臉旁邊（業配感太重）"),
    ]),
    ("各商品最佳拍攝情境", C_VINEGAR, [
        ("味噌", "湯鍋冒煙、木湯匙攪入 → 蒸氣、湯色、豆腐漂浮感"),
        ("味醂", "照燒醬汁收汁、油亮光澤 → 食材反光、醬汁濃稠感"),
        ("烏醋", "炒鍋翻炒、醋液倒入瞬間 → 動態感、高麗菜微焦邊"),
        ("米醋", "涼拌碗、冰鎮後的水珠 → 清爽感、綠色食材"),
        ("辣椒醬", "蛤蜊開口瞬間、拌麵油光 → 紅色醬汁、食慾感"),
    ]),
    ("Reels / 短影片結構", C_RICE_V, [
        ("片長建議", "15–30 秒最佳，超過 60 秒留存率掉快"),
        ("0–3 秒", "成品畫面先出（吊胃口，最重要）"),
        ("3–20 秒", "快剪烹飪過程（只留重點步驟）"),
        ("20–25 秒", "成品特寫"),
        ("25–30 秒", "商品自然入鏡 + 口頭 CTA（「用的是十全的…」）"),
        ("字幕", "加上食材份量、重點步驟文字 → 增加存取率"),
        ("音樂", "找當下 IG 流行音樂，不要用罐頭音效"),
    ]),
    ("不需要準備的東西", C_CHILI, [
        ("不需要", "品牌 Logo 貼片（太廣告感）\n統一的片頭片尾\n特定貼文版型"),
        ("保留 KOL 風格", "我們要的是「真實使用感」，不是「廣告片」。\nKOL 自己的風格才是觸及粉絲的關鍵，不要壓縮它。"),
    ]),
]

r = 2
for section_title, bg, items in sections:
    merge_hdr(ws5, r, 1, r, 4, f"▍ {section_title}", C_SUB, size=10)
    ws5.row_dimensions[r].height = 22
    r += 1
    for label, content in items:
        hdr(ws5, r, 1, label, bg=bg, size=10, center=False)
        cell = ws5.cell(row=r, column=2, value=content)
        cell.fill = PatternFill("solid", fgColor=C_WHITE)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws5.row_dimensions[r].height = max(60, content.count("\n") * 20 + 40)
        r += 1

set_col(ws5, {"A":20, "B":30, "C":20, "D":20})

# ════════════════════════════════════════
# Sheet 6：成效回報表
# ════════════════════════════════════════
ws6 = wb.create_sheet("⑥成效回報表")
merge_hdr(ws6, 1, 1, 1, 7, "KOL 成效回報表（發文後第 7 天填寫回傳）", C_HEADER, size=13)
ws6.row_dimensions[1].height = 30

for ci, h in enumerate(["KOL 帳號","主打商品","發文日期","觸及數","影片播放數","折扣碼使用次數","連結點擊數（UTM）"], 1):
    hdr(ws6, 2, ci, h, bg=C_SUB, size=10)
ws6.row_dimensions[2].height = 22

for i in range(3, 35):
    bg = C_LIGHT if i%2==0 else C_WHITE
    for ci in range(1, 8):
        cell = ws6.cell(row=i, column=ci, value="")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
    ws6.row_dimensions[i].height = 22

# 說明列
r = 36
merge_hdr(ws6, r, 1, r, 7, "📌 說明", C_SUB, size=10)
notes = [
    "折扣碼使用次數：由十全官網後台提供，KOL 不需自行統計",
    "連結點擊數（UTM）：由十全 GA4 後台提供，對照各 KOL 專屬 UTM 參數",
    "觸及數 / 播放數：請 KOL 截圖後台洞察數據，或直接填入數字",
    "回傳方式：填寫完整後 Email 或 LINE 回傳給合作窗口",
]
for i, note in enumerate(notes, r+1):
    cell = ws6.cell(row=i, column=1, value=f"• {note}")
    cell.font = Font(size=10)
    cell.fill = PatternFill("solid", fgColor=C_LIGHT)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border
    ws6.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    ws6.row_dimensions[i].height = 22

set_col(ws6, {"A":20, "B":12, "C":14, "D":14, "E":16, "F":18, "G":20})
ws6.freeze_panes = "A3"

# ════════════════════════════════════════
# 輸出
# ════════════════════════════════════════
output = "/home/user/-/十全特好_KOL素材包範本.xlsx"
wb.save(output)
print(f"✅ {output}")
print("   Sheet 1：KOL 個人資料卡（可複製填寫）")
print("   Sheet 2：商品賣點卡（5 種商品 × 4 個賣點）")
print("   Sheet 3：食譜建議（5 種商品 × 3 道 = 15 道食譜）")
print("   Sheet 4：CTA 話術 + 品牌語調規範")
print("   Sheet 5：拍攝視覺指引（Olivia）")
print("   Sheet 6：成效回報表（發文後第 7 天填寫）")
