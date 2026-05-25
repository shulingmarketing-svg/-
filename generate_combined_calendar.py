import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 顏色 ──
C_HEADER      = "1B3A6B"
C_FB_JUL      = "D6E4F0"
C_FB_AUG      = "D5E8D4"
C_FB_SEP      = "FFF2CC"
C_LINE_JUL    = "FFE0CC"
C_SHOPEE_JUL  = "FDDCCC"
C_SHOPEE_AUG  = "FABBA0"
C_SHOPEE_SEP  = "F79A80"
C_MOMO_JUL    = "EDD5F5"
C_MOMO_AUG    = "D9B0EE"
C_MOMO_SEP    = "C48DE6"
C_LINE_AUG = "FFD0B0"
C_LINE_SEP = "FFBC94"
C_AUTO     = "F0F0F0"
C_EVENT    = "F4CCCC"
C_RECIPE   = "FCE5CD"
C_KNOW     = "EAD1DC"
C_OLIVIA   = "E8D5F5"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(cell, bg, bold=False, center=False, size=10):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=size)
    cell.border = border
    h = "center" if center else "left"
    cell.alignment = Alignment(horizontal=h, vertical="top", wrap_text=True)

# ════════════════════════════════════════
# 資料：FB 24 篇
# ════════════════════════════════════════
fb_posts = [
    ("2025/7/1",  "二", "FB", "活動", "抽獎倒數 2 週，加入會員還來得及", "全系列",
     "6/15 開始的抽獎活動，7/15 第一次開抽。\n\n現在還有兩週，還沒加入官網會員的，\n現在去註冊，還來得及。\n\n加入就送 88 元折扣券，\n折扣券直接抵扣七月限定組合。\n抽獎資格 + 現省 88 元，兩件事一次搞定。\n\n👉 加入連結在留言區",
     "#十全特好 #會員限定 #抽獎活動", "獎品實物 + 倒數「14 天」字樣",
     "朱紅底，白字「倒數 14 天」佔畫面 50%，獎品實物放右側", "大數字、獎品照", "早上 10:00", "連結點擊數、留言數"),
    ("2025/7/4",  "五", "FB", "食譜", "夏天一鍋涼拌，米醋是關鍵", "米醋",
     "不開火也能吃得很好。\n\n小黃瓜拍碎，加蒜末、辣椒、一匙米醋、少許鹽。\n拌一拌，冰 20 分鐘。\n\n米醋不像白醋那麼嗆，酸味圓一點，\n涼拌的時候不會搶過食材的味道。\n\n這道夏天每週至少做兩次，你試試。",
     "#夏日食譜 #涼拌 #十全米醋", "成品俯拍，自然光，不要太精緻感",
     "小黃瓜涼拌俯拍，自然光，米白碗，木桌背景，左上角小字「5 分鐘料理」", "清爽感、夏天色調", "晚上 7:30", "儲存數、分享數"),
    ("2025/7/8",  "二", "FB", "知識", "你家的烏醋是釀的還是調的？", "烏醋",
     "翻一下你家烏醋的成分表。\n\n第幾行出現「焦糖色素」？\n第幾行出現「調味劑」？\n\n十全烏醋：釀造醋、蔬果、香辛料。\n顏色深是因為原料，不是因為色素。\n\n不是每瓶烏醋都一樣，只是大部分人沒有翻過成分表。",
     "#成分控 #天然釀造 #十全烏醋", "兩瓶烏醋成分表並排對比，文字清晰可讀",
     "兩瓶醋並排，成分表各自清楚可讀，深青綠底條做區隔", "對比感強、文字清晰", "早上 10:00", "儲存數、分享數"),
    ("2025/7/11", "五", "FB", "活動", "7/15 開抽，倒數 4 天", "全系列",
     "7/15 開抽，倒數 4 天。\n\n這次抽 300 名，名額比你想的多，\n但沒有加入官網會員就直接出局。\n\n88 元折扣券現在領，\n限定組合現在看，\n7/15 以後就沒有了。\n\n👉 加入連結在留言區",
     "#十全特好 #倒數 #抽獎最後機會", "「倒數 4 天」大字設計，搭配獎品圖",
     "朱紅底，「倒數 4 天」極大白字，下方一行小字說明資格", "緊迫感、數字要大", "晚上 7:30", "連結點擊數、留言數"),
    ("2025/7/15", "二", "FB", "活動", "第一次抽獎得獎公告，300 名出爐", "全系列",
     "第一批 300 位，出爐了。\n\n感謝這個月加入的每一位新會員，\n不管有沒有抽到，88 元折扣券還在，\n下次購買直接用。\n\n第二次抽獎：8/15\n還沒加入的，現在加入還有兩次機會。\n\n得獎名單查詢：[連結]",
     "#十全特好 #抽獎結果 #恭喜得獎", "「恭喜 300 位得獎」簡潔圖卡，品牌色系",
     "深醬色底，「恭喜 300 位」金色字，溫暖收穫感", "慶祝感、金色點綴", "早上 10:00", "留言數、分享數"),
    ("2025/7/18", "五", "FB", "食譜", "味噌湯不只一種喝法，三個懶人版本", "味噌",
     "味噌湯不只一種喝法。\n\n最懶版：味噌 + 豆腐 + 蔥花，三分鐘。\n進階版：加蛤蜊，湯頭甜一個層次。\n夏天版：冷味噌湯，日本人夏天這樣喝。\n\n冷味噌湯做法：\n高湯放涼，加味噌攪開，冰進冰箱，\n加小黃瓜片、冰塊，喝之前再拌一次。\n\n比想像中好喝，你試試。",
     "#味噌湯 #日式料理 #十全味噌", "三種版本的味噌湯並排",
     "三碗味噌湯並排俯拍，分別標「3 分鐘」「進階版」「冷湯」", "清晰對比、三等分構圖", "晚上 7:30", "儲存數、留言數"),
    ("2025/7/22", "二", "FB", "知識", "味醂不是米酒，差在哪裡", "味醂",
     "很多人把味醂當米酒用，錯了。\n\n米酒：去腥，讓肉軟一點。\n味醂：去腥 + 增甜 + 讓食材有光澤。\n\n日式照燒、紅燒、滷味，那個油亮亮的感覺，是味醂的功勞，\n不是糖，不是醬油。\n\n下次滷肉的時候，試著加一匙，\n看看顏色和味道差在哪。",
     "#料理知識 #味醂 #十全味醂", "照燒肉特寫，油亮光澤感明顯",
     "照燒豬肉特寫，油亮光澤，右側文字框說明味醂功效", "食材光澤、文字輔助", "早上 10:00", "儲存數、留言數"),
    ("2025/7/25", "五", "FB", "活動", "8 月抽獎預告，資格自動保留到 9 月", "全系列",
     "7 月抽完了，8 月繼續。\n\n8/15 第二次抽獎，還是 300 名，\n還是官網會員限定。\n\n7 月才加入的也算，\n不用重複加入，資格自動保留到 9 月。\n\n第三次也是最後一次：9/15。\n三個月，共抽 900 名。\n\n還沒加入的，現在是第二次機會。",
     "#十全特好 #抽獎活動 #會員限定", "三個月抽獎時間軸圖卡",
     "三格時間軸圖卡（7/15 ✓ 綠勾、8/15 圓圈、9/15 圓圈），朱紅強調下兩場", "進度感、視覺清楚", "晚上 7:30", "連結點擊數、分享數"),
    ("2025/8/1",  "五", "FB", "食譜", "中元普渡三道家常菜，食材不浪費", "味噌、烏醋、辣椒醬",
     "中元普渡快到了，拜拜完的食材怎麼辦？\n\n三道不浪費的家常菜：\n\n1. 味噌蒸魚\n魚 + 味噌 + 薑絲，電鍋一按，12 分鐘。\n\n2. 醋溜高麗菜\n高麗菜 + 烏醋 + 蒜，大火快炒，酸甜下飯。\n\n3. 辣炒豬肉\n豬肉片 + 辣椒醬 + 蒜，三分鐘，比外送快。\n\n拜完直接煮，不用另外買食材。",
     "#中元節 #普渡料理 #十全調味", "三道菜並排，家常感",
     "三道菜三格構圖，每格一道，左上角各標菜名，家常實物感", "三等分、豐盛感", "晚上 7:30", "儲存數、分享數"),
    ("2025/8/5",  "二", "FB", "活動", "中元禮盒開賣，會員早鳥優惠", "禮盒",
     "中元節送禮，送什麼不失禮？\n\n健康調味禮盒，對象是全家都能用的：\n長輩煮菜用、自己下廚用、送給愛料理的朋友都合適。\n\n十全中元限定禮盒，只在官網賣。\n會員早鳥優惠，本週訂最划算。\n\n送禮不用想太久，\n天然的東西，誰收到都沒有負擔。\n\n👉 禮盒詳情在留言區",
     "#中元送禮 #十全特好 #健康禮盒", "禮盒實物正面，質感包裝",
     "禮盒正面大圖，乾淨白底，右側文字「會員早鳥」紅色標籤", "產品主角、標籤搶眼", "早上 10:00", "連結點擊數、私訊數"),
    ("2025/8/8",  "五", "FB", "知識", "辣椒醬成分表，你上次看是什麼時候", "辣椒醬",
     "辣椒醬的成分表，你上次看是什麼時候？\n\n常見的：辣椒、大豆油、防腐劑、調味劑、色素。\n十全的：辣椒、蒜、鹽、少量釀造醋。\n\n辣的來源不一樣，\n合成辣和天然辣在嘴裡的感覺不一樣，\n吃完喉嚨的感受也不一樣。\n\n試過天然辣椒醬的人，\n很少回頭買合成的。",
     "#成分控 #辣椒醬 #十全辣椒醬", "成分表對比圖，文字要夠大夠清楚",
     "辣椒醬瓶身 + 成分表放大，兩欄對比，左合成右天然", "對比版面、文字大", "晚上 7:30", "儲存數、分享數"),
    ("2025/8/12", "二", "FB", "活動", "8/15 第二次抽獎，倒數 3 天", "全系列",
     "8/15 第二次抽獎，倒數 3 天。\n\n7 月已經抽了 300 名，\n這次再抽 300 名，\n9/15 最後一次，還有 300 名。\n\n現在加入會員，\n後面兩次都有機會。\n\n88 元折扣券加入就送，資格不會過期。\n\n👉 官網加入連結在留言區",
     "#十全特好 #抽獎倒數 #會員限定", "「倒數 3 天」+ 累計已抽 300 名圖卡",
     "朱紅底，「倒數 3 天」白字大，下方進度條「600/900」", "緊迫感、累積進度", "早上 10:00", "連結點擊數、留言數"),
    ("2025/8/15", "五", "FB", "活動", "第二次抽獎公告，累計 600 名得獎", "全系列",
     "第二批 300 位，出爐了。\n\n兩個月，累計 600 位得獎會員。\n\n還沒抽到的，9/15 還有最後一次機會，\n300 名，官網會員限定，資格自動保留。\n\n還沒加入的，\n現在加入，9/15 還來得及。\n\n得獎名單查詢：[連結]",
     "#十全特好 #抽獎結果 #恭喜得獎", "「累計 600 位得獎」圖卡，有進度感",
     "深醬色底，「累計 600 位得獎」金字，中間進度條已填 2/3", "進度感、溫暖", "早上 10:00", "留言數、分享數"),
    ("2025/8/19", "二", "FB", "食譜", "辣椒醬不只沾著吃，三種懶人料理", "辣椒醬",
     "辣椒醬不只是沾醬。\n\n三種懶人用法：\n\n1. 辣炒蛤蜊\n蛤蜊 + 辣椒醬 + 薑，大火兩分鐘，\n比餐廳的還香。\n\n2. 拌麵\n煮好的麵 + 辣椒醬 + 少許麻油 + 蔥花，\n五分鐘一碗，深夜救星。\n\n3. 醃肉\n雞腿 + 辣椒醬 + 蒜末，冰一晚，\n隔天烤或煎都好吃。\n\n一瓶用三種方式，物超所值。",
     "#辣椒醬料理 #懶人食譜 #十全辣椒醬", "三種料理並排",
     "三種料理橫排或三格，蛤蜊/拌麵/烤雞各一，自然光", "豐富感、料理質感", "晚上 7:30", "儲存數、留言數"),
    ("2025/8/22", "五", "FB", "知識", "釀造醋 vs 調製醋，成分表不會說謊", "烏醋、米醋",
     "買醋之前，先翻成分表。\n\n調製醋：水、醋酸、焦糖色素、調味劑。\n釀造醋：糧食原料、時間。\n\n兩種都叫醋，價格差不多，\n進你身體的東西差很多。\n\n不是叫你一定買我們的，\n是叫你養成翻成分表的習慣。",
     "#成分控 #釀造醋 #十全烏醋", "兩種成分表並排對比",
     "成分表兩欄並排，左側打 ✗，右側打 ✓，深青綠底強調", "對比清晰、判斷感", "晚上 7:30", "儲存數、分享數"),
    ("2025/8/26", "二", "FB", "活動", "9/15 最終場預告 + 中秋禮盒預熱", "全系列、禮盒",
     "9/15 是最後一次，第三次抽獎。\n\n三個月抽了 600 名，\n最後這批 300 名，\n9/15 壓軸。\n\n同時，中秋禮盒下週開賣。\n送禮的對象，你心裡有數了嗎？\n\n天然調味禮盒，給在乎吃什麼的人。\n\n詳情下週公布，先追蹤不要錯過。",
     "#十全特好 #中秋禮盒 #抽獎最終場", "一半中秋氛圍、一半抽獎倒數",
     "上半中秋月亮氛圍，下半「最終場 9/15」倒數，分割構圖", "雙主題、視覺切割", "早上 10:00", "追蹤數、分享數"),
    ("2025/9/2",  "二", "FB", "活動", "中秋禮盒正式開賣，會員優先購", "禮盒",
     "中秋送禮，今年換個思路。\n\n月餅吃完就忘，\n調味料用一整年。\n\n十全中秋限定禮盒：\n味噌、味醂、烏醋、米醋，四件組。\n天然釀造，成分看得懂。\n\n收到的人會記得你送的是什麼。\n\n官網會員優先購，今天開賣。\n\n👉 禮盒詳情在留言區",
     "#中秋送禮 #十全特好 #天然調味禮盒", "禮盒與中秋氛圍結合",
     "禮盒 + 月亮背景，深夜質感，右側「會員優先購」標籤", "中秋氛圍、禮品感", "早上 10:00", "連結點擊數、私訊數"),
    ("2025/9/5",  "五", "FB", "食譜", "中秋烤肉醬自己調，烏醋 + 辣椒醬", "烏醋、辣椒醬",
     "中秋烤肉，醬料自己調的和買的，差一個層次。\n\n基底烤肉醬配方：\n醬油 3 匙 + 烏醋 1 匙 + 辣椒醬半匙 + 蒜末 + 少許糖\n\n烏醋讓醬料多一個發酵的層次，\n辣椒醬帶微辣不嗆，\n蒜末烤過之後香氣全出來。\n\n做一碗大家沾，\n比瓶裝烤肉醬好吃，成本少很多。",
     "#中秋烤肉 #烤肉醬 #十全調味", "烤肉實境，搭配醬料小碗",
     "烤肉現場俯拍，炭火煙霧感，小碗自調醬料放前景", "煙霧感、生活感強", "晚上 7:30", "儲存數、分享數"),
    ("2025/9/9",  "二", "FB", "活動", "最終場抽獎倒數 6 天，這次沒加入就真的沒了", "全系列",
     "9/15，最後一次。\n\n三個月，共三次抽獎，每次 300 名。\n最後這場，還沒抽到的還有機會。\n\n現在加入官網會員，\n9/15 還來得及。\n\n加入就送 88 元折扣券，\n中秋禮盒折扣直接用。\n\n倒數 6 天，\n這次沒加入，就真的沒了。\n\n👉 加入連結在留言區",
     "#十全特好 #抽獎最終場 #倒數", "「最終場 倒數 6 天」大字圖卡",
     "朱紅底，「最終場 倒數 6 天」白字極大，緊張感拉滿", "最大衝擊感", "早上 10:00", "連結點擊數、留言數"),
    ("2025/9/12", "五", "FB", "知識", "好的味噌，需要時間", "味噌",
     "好的味噌，需要時間。\n\n黃豆、米麴、鹽，\n裝進木桶，壓上重石，\n等三個月，或者更久。\n\n這段時間裡什麼都不做，\n只是讓菌慢慢把黃豆變成味噌。\n\n快速製程的味噌也有，\n味道不一樣，用途不一樣，\n你的舌頭會告訴你差在哪。",
     "#味噌 #發酵食品 #十全味噌", "釀造過程質感照",
     "木桶或陶甕，靜物感，柔和打光，日系工藝質感", "時間感、工藝感", "晚上 7:30", "儲存數、留言數"),
    ("2025/9/15", "一", "FB", "活動", "最終場抽獎公告，累計 900 名得獎，感謝支持", "全系列",
     "最後一批，300 位，出爐了。\n\n三個月，三次抽獎，\n累計 900 位得獎會員，謝謝大家。\n\n沒有抽到的，\n你的 88 元折扣券還在，\n會員資格也還在，\n我們還有很多好東西等你。\n\n得獎名單查詢：[連結]\n\n下一個活動，等我們。",
     "#十全特好 #抽獎結果 #感謝支持", "「累計 900 位得獎，感謝支持」溫暖收尾圖卡",
     "深醬色底，「累計 900 位，謝謝你們」金字，溫暖收尾", "感謝感、完結感", "早上 10:00", "留言數、分享數"),
    ("2025/9/19", "五", "FB", "食譜", "秋天一鍋味噌豬肉湯，周末煮一次", "味噌",
     "天氣一涼，就想煮這鍋。\n\n味噌豬肉湯，周末煮一次，\n家裡會很香。\n\n食材：豬肉片、白蘿蔔、豆腐、蔥\n\n做法：\n蘿蔔先煮軟，加豬肉片燙熟，\n關小火，味噌用篩網濾入湯裡，\n不要再煮滾，味噌的香氣才不會跑掉。\n蔥花最後撒，起鍋。\n\n這鍋從秋天吃到冬天，\n你會一直煮。",
     "#秋天食譜 #味噌豬肉湯 #十全味噌", "熱湯冒煙的溫暖感，秋天色調",
     "味噌豬肉湯，蒸氣冒煙，秋天橘調光線，木碗或陶碗", "暖色、食慾感", "晚上 7:30", "儲存數、留言數"),
    ("2025/9/23", "二", "FB", "活動", "中秋禮盒最後機會 + 雙十會員搶先看預告", "禮盒、全系列",
     "中秋禮盒，最後幾天。\n\n還沒送的，今天訂，還來得及。\n\n另外先說一聲：\n雙十連假，十全官網會員有特別的東西，\n細節下週公布。\n\n還沒加入會員的，\n這週加入，雙十活動你才看得到。\n\n👉 中秋禮盒連結在留言區",
     "#中秋禮盒 #十全特好 #雙十預告", "中秋收尾 + 雙十倒數感",
     "左禮盒、右「雙十預告」箭頭指向，分割畫面設計", "兩個訊息並陳", "早上 10:00", "連結點擊數、追蹤數"),
    ("2025/9/26", "五", "FB", "知識", "天然調味料怎麼挑？看這三個字就夠了", "全系列",
     "天然調味料怎麼挑？\n看這三個字就夠了。\n\n「釀造」。\n\n醬油有釀造醬油，醋有釀造醋，\n釀造的意思是：原料 + 時間，沒有捷徑。\n\n快速製程也有產品，\n味道可以調得很像，\n但製程不一樣，進身體的東西就不一樣。\n\n買之前翻一下成分表，\n「釀造」兩個字在不在，答案就出來了。",
     "#天然調味 #釀造 #十全特好", "成分表特寫，「釀造」兩個字清楚圈出",
     "成分表特寫，紅圈圈出「釀造」兩字，深青綠底色", "聚焦細節、強調字", "晚上 7:30", "儲存數、分享數"),
]

# ════════════════════════════════════════
# 資料：LINE OA 13 則推播
# ════════════════════════════════════════
line_posts = [
    ("2025/7/1",  "二", "LINE OA", "喚醒推播", "舊好友喚醒：一鍵加入官網會員",
     "你加入我們 LINE 已經一段時間了。\n\n但 88 元折扣券你拿到了嗎？\n\n用 LINE 帳號一鍵加入官網會員，\n券馬上匯入，7/15 抽獎資格也一起有。\n\n這次沒加入，就真的沒了。\n👇 點這裡，10 秒完成",
     "早上 10:00", "連結點擊數、加入會員數"),
    ("2025/7/11", "五", "LINE OA", "活動倒數", "7/15 開抽，倒數 4 天",
     "7/15 開抽，還有 4 天。\n\n資格確認方法：登入官網，有帳號就 OK。\n沒有帳號的，現在還來得及，用 LINE 一鍵就好。\n\n👇 確認資格",
     "晚上 7:30", "連結點擊數"),
    ("2025/7/15", "二", "LINE OA", "得獎公告", "第一次抽獎 300 名得獎公告",
     "第一批 300 位，出來了。\n\n得獎名單 → [連結]\n\n沒抽到也沒關係，88 元券還在帳號裡，\n下次買直接用。8/15 還有一次。",
     "早上 10:00", "連結點擊數、留言數"),
    ("2025/7/25", "五", "LINE OA", "活動預告", "8 月抽獎 + 中元禮盒預熱",
     "下週開始，我們有兩件事要跟你說：\n\n8/15 第二次抽獎，還有 300 名。\n中元禮盒下週開賣，LINE 好友早鳥優先。\n\n都不用特別做什麼，繼續待著就好。",
     "晚上 7:30", "訊息開啟率"),
    ("2025/8/5",  "二", "LINE OA", "禮盒限定", "中元禮盒 LINE 好友早鳥（比 FB 早一天）",
     "中元禮盒今天開賣，LINE 好友專屬早鳥價。\n\nFB 明天才公告，你早一天看到。\n庫存有限，今天訂比較保險。\n\n👇 禮盒頁面",
     "早上 10:00", "連結點擊數、訂單數"),
    ("2025/8/12", "二", "LINE OA", "活動倒數", "8/15 開抽，倒數 3 天",
     "8/15 第二次抽獎，倒數 3 天。\n\n7 月已抽了 300 名，這次再抽 300 名。\n還沒成為官網會員的，今天是最後機會。\n\n👇 用 LINE 一鍵加入",
     "早上 10:00", "連結點擊數"),
    ("2025/8/15", "五", "LINE OA", "得獎公告", "第二次抽獎 300 名得獎公告，累計 600 名",
     "第二批 300 位，出爐了。\n\n兩個月累計 600 位得獎。\n得獎名單 → [連結]\n\n9/15 最後一次，300 名，資格自動保留。",
     "早上 10:00", "連結點擊數、留言數"),
    ("2025/8/22", "五", "LINE OA", "食譜推播", "本月最多人存的食譜：辣椒醬三種懶人料理",
     "這個月 FB 最多人存的一篇：\n「辣椒醬三種懶人料理」\n\n拌麵、炒蛤蜊、醃雞腿，你試過哪一種？\n👇 看完整食譜",
     "晚上 7:30", "連結點擊數、回覆數"),
    ("2025/9/2",  "二", "LINE OA", "禮盒限定", "中秋禮盒 LINE 好友早鳥（比 FB 早一天）",
     "中秋禮盒今天開賣。\nLINE 好友專屬早鳥連結在下面，比 FB 早一天。\n\n味噌、味醂、烏醋、米醋四件組，送禮用一整年。\n👇",
     "早上 10:00", "連結點擊數、訂單數"),
    ("2025/9/9",  "二", "LINE OA", "活動倒數", "最終場 9/15，倒數 6 天",
     "9/15 最後一次抽獎，倒數 6 天。\n\n三個月共抽 900 名，最後這批 300 名。\n還沒有官網帳號的，現在用 LINE 一鍵加入。\n\n這次沒加入，就真的沒了。\n👇",
     "早上 10:00", "連結點擊數"),
    ("2025/9/15", "一", "LINE OA", "得獎公告", "最終場 300 名得獎公告，累計 900 名感謝",
     "最後一批 300 位，出爐了。\n\n三個月，900 位得獎，謝謝你們。\n得獎名單 → [連結]\n\n沒抽到的，88 元券還在，我們還有好東西。\n10 月見。",
     "早上 10:00", "連結點擊數、留言數"),
    ("2025/9/23", "二", "LINE OA", "活動預告", "雙十會員搶先看預告",
     "雙十連假，十全有東西要給你。\n\n細節下週公布，官網會員才看得到。\n還沒加入的，這週加入來得及。\n\n👇",
     "早上 10:00", "連結點擊數、加入會員數"),
    ("2025/9/30", "二", "LINE OA", "暖場", "10 月有好東西，先追蹤不要錯過",
     "十月，我們有好東西要給你。\n\n細節下週公布，官網會員才看得到。\n還沒加入的，現在加入剛好。\n\n👇 用 LINE 一鍵加入會員",
     "晚上 7:30", "訊息開啟率"),
]

# LINE OA 自動序列
line_auto = [
    ("自動觸發", "加入當下", "新好友加入：立即歡迎 + 一鍵入會",
     "嗨，歡迎加入！\n\n你已經在 LINE 了，官網會員一鍵就好。\n點下面用 LINE 帳號登入，88 元折扣券馬上進帳。\n7/15 抽獎資格也一起有，不用再做別的。\n\n👇 一鍵加入官網會員"),
    ("自動觸發", "加入後第 2 天", "新好友未轉換：低摩擦再提醒",
     "折扣券還沒領到？\n\n不用填表單，不用記密碼。\n用你的 LINE 帳號點一下就完成了。\n\n👇 10 秒搞定"),
    ("自動觸發", "加入後第 5 天", "新好友未轉換：最後提醒 + 抽獎緊迫感",
     "最後提醒一次。\n\n7/15 抽獎只有官網會員才能參加，\n現在加入還來得及。\n\n88 元折扣券、抽獎資格，一起有。\n👇"),
]

# ════════════════════════════════════════
# 建立 Excel
# ════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet 1：綜合排程月曆 ──
ws1 = wb.active
ws1.title = "📅 綜合排程月曆"

combo_headers = ["日期", "星期", "渠道", "類型", "主題／摘要", "完整內文", "素材／備注", "發送時間", "追蹤指標", "✅ 完成"]
combo_widths  = [12, 6, 10, 10, 35, 55, 30, 12, 20, 8]

for ci, (h, w) in enumerate(zip(combo_headers, combo_widths), 1):
    cell = ws1.cell(row=1, column=ci, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.row_dimensions[1].height = 22

# ════════════════════════════════════════
# 資料：蝦皮活動
# ════════════════════════════════════════
shopee_posts = [
    ("2025/7/1",  "二", "蝦皮", "活動開跑",
     "十全好食節開跑｜任 2 件 9 折 + 蝦幣回饋",
     "【活動期間】7/1–7/25\n【主打商品】即飲果醋 + 蒟蒻飲組合\n【優惠】任選 2 件 9 折 + 蝦幣回饋\n【配合動作】訂單包裝內夾官網入會 QR Code 卡",
     "上架前確認商品頁圖片與活動標籤設定", "上午 10:00", "訂單數、新客比例"),
    ("2025/7/25", "五", "蝦皮", "活動收尾",
     "好食節收尾倒數｜8 月中元預告",
     "【動作】更新商品頁 Banner 預告 8 月中元活動\n【文案方向】「好食節感謝，中元好料即將登場」\n【注意】確保庫存備貨到位",
     "更新蝦皮店頭 Banner", "下午 2:00", "頁面瀏覽數"),
    ("2025/8/10", "日", "蝦皮", "活動預告",
     "中元閃購倒數預告｜8/12–8/14 每天 2 小時限時",
     "【預告內容】8/12、8/13、8/14 每天 14:00–16:00 閃購\n【主打商品】味噌 + 烏醋 + 辣椒醬三件組\n【預告動作】限時動態 + 商品頁倒數標籤",
     "設定蝦皮閃購時段，確認庫存", "上午 10:00", "加購清單數"),
    ("2025/8/12", "二", "蝦皮", "限時閃購",
     "中元閃購 Day 1｜14:00–16:00",
     "【時段】14:00–16:00\n【折扣】限時 85 折\n【商品】味噌 + 烏醋 + 辣椒醬三件組\n【配合】同步在 LINE OA 提醒好友",
     "閃購期間監控庫存與訂單狀況", "14:00", "閃購訂單數、GMV"),
    ("2025/8/13", "三", "蝦皮", "限時閃購",
     "中元閃購 Day 2｜14:00–16:00",
     "【時段】14:00–16:00\n【折扣】限時 85 折\n【商品】同 Day 1\n【加碼】留評價送小禮（若庫存允許）",
     "確認 Day 1 庫存餘量，調整 Day 2 數量", "14:00", "閃購訂單數、評價數"),
    ("2025/8/14", "四", "蝦皮", "限時閃購",
     "中元閃購 Day 3（最終場）｜14:00–16:00",
     "【時段】14:00–16:00\n【折扣】限時 85 折\n【文案】「今天最後一場，明天恢復原價」\n【配合】FB + LINE OA 最後提醒",
     "最後一天加強社群提醒力道", "14:00", "閃購訂單數、三天累計 GMV"),
    ("2025/9/8",  "一", "蝦皮", "早鳥預購",
     "中秋早鳥預購開始｜9/8–9/10 限定 9 折",
     "【活動期間】9/8–9/10\n【折扣】早鳥 9 折\n【主打商品】味噌 + 味醂 + 烏醋 + 米醋四件組\n【加碼】前 100 組附「烤肉醬食譜卡」\n【定價原則】不低於官網會員價",
     "設定早鳥折扣碼，確認食譜卡印製數量", "上午 10:00", "預購訂單數"),
    ("2025/9/12", "五", "蝦皮", "正式開賣",
     "中秋禮盒正式開賣（與官網同步）",
     "【動作】早鳥結束，恢復原價開賣\n【配合】FB + LINE OA 同步公告\n【文案】「中秋送禮，天然調味料用一整年」\n【注意】確保物流在中秋前 3 天可到貨",
     "確認物流時效，中秋前 3 天截單", "上午 10:00", "中秋期間總訂單數、GMV"),
]

# ════════════════════════════════════════
# 資料：MOMO 活動
# ════════════════════════════════════════
momo_posts = [
    ("2025/8/10", "日", "MOMO", "商品上架",
     "中元禮盒上架｜滿 1,200 免運 + 贈品",
     "【商品】中元天然調味禮盒（味噌 + 烏醋 + 辣椒醬）\n【定價】與官網同步\n【優惠】滿 1,200 免運 + 贈品（小包裝試用品）\n【圖片】使用官網禮盒實物圖",
     "確認 MOMO 商品頁審核通過，圖片規格符合平台要求", "上午 10:00", "商品頁瀏覽數、加購數"),
    ("2025/8/15", "五", "MOMO", "節慶衝單",
     "中元禮盒主力日｜配合抽獎公告衝單",
     "【動作】配合 FB + LINE 抽獎公告，在 MOMO 同步推廣\n【文案方向】「送禮送健康，天然調味全家用」\n【可申請】MOMO 節慶聯合檔期（若有補貼）",
     "查詢 MOMO 8 月是否有聯合檔期補貼可申請", "上午 10:00", "當日訂單數、禮盒銷售佔比"),
    ("2025/9/8",  "一", "MOMO", "商品上架",
     "中秋禮盒上架｜滿 2,000 加購優惠",
     "【商品】中秋天然調味禮盒（味噌 + 味醂 + 烏醋 + 米醋四件組）\n【優惠】滿 2,000 加購第二件 75 折\n【目標客群】30–55 歲，送禮給長輩或朋友",
     "確認 MOMO 商品頁上架，主圖需有中秋氛圍", "上午 10:00", "商品頁瀏覽數、加購數"),
    ("2025/9/12", "五", "MOMO", "節慶衝單",
     "中秋正式開賣｜主力推廣日",
     "【動作】與蝦皮、官網同步開賣\n【MOMO 特色】主打禮盒送禮情境，文案偏「送給爸媽」\n【可申請】MOMO 中秋聯合檔期（若有補貼）",
     "查詢 MOMO 9 月中秋聯合檔期補貼，截止申請日期", "上午 10:00", "當日訂單數、GMV"),
    ("2025/9/17", "三", "MOMO", "活動收尾",
     "中秋禮盒最後衝刺｜中秋前 2 天截單",
     "【動作】更新商品頁「中秋前 2 天下單才來得及」緊迫文案\n【配合】FB + LINE 同步提醒\n【截單時間】9/17 23:59",
     "確認物流可在 9/19 前到貨，超過截止下架禮盒", "上午 10:00", "最終累計訂單數、中秋期間總 GMV"),
]

# 合併 FB + LINE + 蝦皮 + MOMO，按日期排序
month_fb     = {"7": C_FB_JUL,      "8": C_FB_AUG,      "9": C_FB_SEP}
month_line   = {"7": C_LINE_JUL,    "8": C_LINE_AUG,    "9": C_LINE_SEP}
month_shopee = {"7": C_SHOPEE_JUL,  "8": C_SHOPEE_AUG,  "9": C_SHOPEE_SEP}
month_momo   = {"7": C_MOMO_JUL,    "8": C_MOMO_AUG,    "9": C_MOMO_SEP}

type_color = {
    "活動": C_EVENT, "食譜": C_RECIPE, "知識": C_KNOW,
    "喚醒推播": "FFAA80", "活動倒數": "FF9966", "得獎公告": "FF8855",
    "禮盒限定": "FF7744", "食譜推播": "FF6633", "活動預告": "FF5522", "暖場": "FF4411",
    "活動開跑": "FFA070", "活動收尾": "FF8060", "限時閃購": "FF6050",
    "早鳥預購": "FF9080", "正式開賣": "FF7060", "活動預告": "FF8050",
    "商品上架": "CC88DD", "節慶衝單": "BB66CC",
}

all_rows = []
for p in fb_posts:
    date, dow, ch, typ, title, prod, body, htag, img, cfg_dir, key_elm, send_t, kpi = p
    note = f"圖片：{img}\nOlivia：{cfg_dir}｜{key_elm}"
    all_rows.append((date, dow, ch, typ, title, body, note, send_t, kpi))

for p in line_posts:
    date, dow, ch, typ, title, body, send_t, kpi = p
    all_rows.append((date, dow, ch, typ, title, body, "", send_t, kpi))

for p in shopee_posts:
    date, dow, ch, typ, title, body, note, send_t, kpi = p
    all_rows.append((date, dow, ch, typ, title, body, note, send_t, kpi))

for p in momo_posts:
    date, dow, ch, typ, title, body, note, send_t, kpi = p
    all_rows.append((date, dow, ch, typ, title, body, note, send_t, kpi))

all_rows.sort(key=lambda x: x[0])

def get_bg(ch, month):
    if ch == "LINE OA": return month_line[month]
    if ch == "蝦皮":    return month_shopee[month]
    if ch == "MOMO":    return month_momo[month]
    return month_fb[month]

for ri, row in enumerate(all_rows, 2):
    date, dow, ch, typ, title, body, note, send_t, kpi = row
    month = date.split("/")[1]
    bg = get_bg(ch, month)
    type_bg = type_color.get(typ, "FFFFFF")
    vals = [date, dow, ch, typ, title, body, note, send_t, kpi, ""]
    for ci, v in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=ci, value=v)
        cell.border = border
        if ci == 4:
            cell.fill = PatternFill("solid", fgColor=type_bg)
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        else:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if ci in (1,2,3,8,10) else "left")
    ws1.row_dimensions[ri].height = 100

ws1.freeze_panes = "A2"

# ── Sheet 2：FB 貼文完整稿 ──
ws2 = wb.create_sheet("📘 FB 貼文完整稿")
fb_headers = ["編號","日期","星期","類型","主題","主打商品","貼文內文","Hashtag","圖片建議","配圖方向（Olivia）","重點元素","發文時間","追蹤指標"]
fb_widths  = [5,12,6,8,30,16,58,28,24,38,18,12,18]

for ci,(h,w) in enumerate(zip(fb_headers,fb_widths),1):
    cell = ws2.cell(row=1,column=ci,value=h)
    bg = "6A1B9A" if ci>=10 else C_HEADER
    cell.font = Font(bold=True,color="FFFFFF",size=10)
    cell.fill = PatternFill("solid",fgColor=bg)
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[1].height = 22

for ri,p in enumerate(fb_posts,2):
    date,dow,ch,typ,title,prod,body,htag,img,cfg_dir,key_elm,send_t,kpi = p
    month = date.split("/")[1]
    row_bg = month_fb[month]
    type_bg = type_color.get(typ,"FFFFFF")
    vals = [ri-1,date,dow,typ,title,prod,body,htag,img,cfg_dir,key_elm,send_t,kpi]
    for ci,v in enumerate(vals,1):
        cell = ws2.cell(row=ri,column=ci,value=v)
        cell.border = border
        if ci==4:
            cell.fill = PatternFill("solid",fgColor=type_bg)
            cell.alignment = Alignment(horizontal="center",vertical="top",wrap_text=True)
        elif ci>=10:
            cell.fill = PatternFill("solid",fgColor=C_OLIVIA)
            cell.alignment = Alignment(vertical="top",wrap_text=True)
        else:
            cell.fill = PatternFill("solid",fgColor=row_bg)
            cell.alignment = Alignment(vertical="top",wrap_text=True,
                                       horizontal="center" if ci in (1,2,3,12) else "left")
    ws2.row_dimensions[ri].height = 110

ws2.freeze_panes = "A2"

# ── Sheet 3：LINE OA 推播完整稿 ──
ws3 = wb.create_sheet("💬 LINE OA 推播完整稿")
line_headers = ["編號","日期","星期","類型","主題","完整訊息內容","發送時間","追蹤指標"]
line_widths  = [5,12,6,12,32,55,12,20]

for ci,(h,w) in enumerate(zip(line_headers,line_widths),1):
    cell = ws3.cell(row=1,column=ci,value=h)
    cell.font = Font(bold=True,color="FFFFFF",size=10)
    cell.fill = PatternFill("solid",fgColor="E65100")
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[1].height = 22

for ri,p in enumerate(line_posts,2):
    date,dow,ch,typ,title,body,send_t,kpi = p
    month = date.split("/")[1]
    bg = month_line[month]
    type_bg = type_color.get(typ,"FFCC99")
    vals = [ri-1,date,dow,typ,title,body,send_t,kpi]
    for ci,v in enumerate(vals,1):
        cell = ws3.cell(row=ri,column=ci,value=v)
        cell.border = border
        if ci==4:
            cell.fill = PatternFill("solid",fgColor=type_bg)
            cell.alignment = Alignment(horizontal="center",vertical="top",wrap_text=True)
        else:
            cell.fill = PatternFill("solid",fgColor=bg)
            cell.alignment = Alignment(vertical="top",wrap_text=True,
                                       horizontal="center" if ci in (1,2,3,7) else "left")
    ws3.row_dimensions[ri].height = 100

ws3.freeze_panes = "A2"

# ── Sheet 4：LINE OA 自動序列 ──
ws4 = wb.create_sheet("🤖 LINE 自動序列")
auto_headers = ["觸發時機","時間點","主題","完整訊息內容"]
auto_widths  = [14,16,30,55]

for ci,(h,w) in enumerate(zip(auto_headers,auto_widths),1):
    cell = ws4.cell(row=1,column=ci,value=h)
    cell.font = Font(bold=True,color="FFFFFF",size=10)
    cell.fill = PatternFill("solid",fgColor="455A64")
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws4.column_dimensions[get_column_letter(ci)].width = w
ws4.row_dimensions[1].height = 22

for ri,p in enumerate(line_auto,2):
    trigger,timing,title,body = p
    vals = [trigger,timing,title,body]
    for ci,v in enumerate(vals,1):
        cell = ws4.cell(row=ri,column=ci,value=v)
        cell.border = border
        cell.fill = PatternFill("solid",fgColor=C_AUTO)
        cell.alignment = Alignment(vertical="top",wrap_text=True,
                                   horizontal="center" if ci in (1,2) else "left")
    ws4.row_dimensions[ri].height = 80

# ── Sheet 5：蝦皮 / MOMO 活動完整稿 ──
ws5 = wb.create_sheet("🛒 蝦皮MOMO活動完整稿")
ec_headers = ["編號","日期","星期","平台","類型","活動主題","活動內容","執行備注","發布時間","追蹤指標"]
ec_widths  = [5,12,6,8,10,30,50,30,12,20]

for ci,(h,w) in enumerate(zip(ec_headers,ec_widths),1):
    cell = ws5.cell(row=1,column=ci,value=h)
    cell.font = Font(bold=True,color="FFFFFF",size=10)
    cell.fill = PatternFill("solid",fgColor="B71C1C")
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws5.column_dimensions[get_column_letter(ci)].width = w
ws5.row_dimensions[1].height = 22

ec_all = [(p, "蝦皮") for p in shopee_posts] + [(p, "MOMO") for p in momo_posts]
ec_all.sort(key=lambda x: x[0][0])

for ri,(p,_) in enumerate(ec_all,2):
    date,dow,ch,typ,title,body,note,send_t,kpi = p
    month = date.split("/")[1]
    bg = month_shopee[month] if ch=="蝦皮" else month_momo[month]
    type_bg = type_color.get(typ,"FFDDCC")
    vals = [ri-1,date,dow,ch,typ,title,body,note,send_t,kpi]
    for ci,v in enumerate(vals,1):
        cell = ws5.cell(row=ri,column=ci,value=v)
        cell.border = border
        if ci in (4,5):
            cell.fill = PatternFill("solid",fgColor=type_bg if ci==5 else
                                    (C_SHOPEE_AUG if ch=="蝦皮" else C_MOMO_AUG))
            cell.alignment = Alignment(horizontal="center",vertical="top",wrap_text=True)
        else:
            cell.fill = PatternFill("solid",fgColor=bg)
            cell.alignment = Alignment(vertical="top",wrap_text=True,
                                       horizontal="center" if ci in (1,2,3,9) else "left")
    ws5.row_dimensions[ri].height = 100

ws5.freeze_panes = "A2"

# ── Sheet 6：圖例 ──
ws6 = wb.create_sheet("📌 圖例說明")
legend = [
    ("顏色","說明"),
    ("淺藍","FB 7 月貼文"),("淺綠","FB 8 月貼文"),("淺黃","FB 9 月貼文"),
    ("淺橘","LINE OA 7 月推播"),("中橘","LINE OA 8 月推播"),("深橘","LINE OA 9 月推播"),
    ("淺珊瑚","蝦皮 7 月"),("中珊瑚","蝦皮 8 月"),("深珊瑚","蝦皮 9 月"),
    ("淺紫","MOMO 7 月"),("中紫","MOMO 8 月"),("深紫","MOMO 9 月"),
    ("",""),
    ("Sheet 說明",""),
    ("📅 綜合排程月曆","FB + LINE OA + 蝦皮 + MOMO 全通路按日期合併"),
    ("📘 FB 貼文完整稿","FB 24 篇含 Olivia 配圖方向"),
    ("💬 LINE OA 推播完整稿","13 則定期推播訊息"),
    ("🤖 LINE 自動序列","新好友加入後的 3 則自動回覆"),
    ("🛒 蝦皮MOMO活動完整稿","蝦皮 8 則 + MOMO 5 則，共 13 筆活動"),
    ("",""),
    ("重要原則",""),
    ("定價","蝦皮 / MOMO 售價 ≥ 官網原價，保護會員制稀缺性"),
    ("時序","官網會員 → LINE OA → 蝦皮 = MOMO → FB，會員永遠最早"),
    ("導流","每筆蝦皮 / MOMO 訂單包裝夾「官網入會 QR Code 卡」"),
]
for ri,(a,b) in enumerate(legend,1):
    bold = ri==1 or ri==14 or ri==21
    ws6.cell(row=ri,column=1,value=a).font = Font(bold=bold)
    ws6.cell(row=ri,column=2,value=b)
ws6.column_dimensions["A"].width = 22
ws6.column_dimensions["B"].width = 50

output = "/home/user/-/十全特好_7-9月完整排程月曆.xlsx"
wb.save(output)
print(f"✅ {output}")
print(f"   FB：{len(fb_posts)} 篇 ｜ LINE OA：{len(line_posts)} 則 ｜ 蝦皮：{len(shopee_posts)} 則 ｜ MOMO：{len(momo_posts)} 則")
print(f"   綜合月曆共 {len(all_rows)} 筆，按日期排序")
